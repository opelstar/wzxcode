from openpyxl import load_workbook


class ReadExcel:

    def __init__(self, file, sheet):
        """
        :param file:需要解析的文件路径
        :param sheet:表名
        """
        self.sheet_name = load_workbook(filename=file)
        self.sheet_values = self.sheet_name[sheet]  # 打开对于的sheet表

    def read_excel(self):
        data = []  # 定义一个列表接收每一组数据
        title = {}  # 定义一个字典
        for i in range(1, self.sheet_values.max_column + 1):
            title[i] = self.sheet_values.cell(1, i).value  # 遍历最大列依次获取每个单元格的value

        for i in range(2, self.sheet_values.max_row + 1):  # 从第二行开始遍历最大行
            values = {}
            for j in range(1, self.sheet_values.max_column + 1):  # 从第二行第一列开始遍历单元格
                # 获取每个单元格的数据后，通过title去匹配标题对应的value数据
                values[title[j]] = self.sheet_values.cell(i, j).value
            data.append(values)  # 把每一组数据添加到data列表内
        # [{key1：value1,key2：value2},{...},...]
        return data


excel_path = 'C:\pythonwork\data\wzx.xlsx'
cases = ReadExcel(excel_path,'Sheet1')
case = cases.read_excel()
print(case)

